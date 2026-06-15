from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from accounts.models import CustomUser
from django.contrib import messages
from django.db.models import Q
from django.contrib.auth.hashers import check_password
from django.views.decorators.http import require_POST
from .models import Appointment, AddedPatient
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors



def is_doctor(user):
    return user.role == "doctor"


@login_required
@user_passes_test(is_doctor)
def doctor_dashboard(request):
    doctor = request.user
    today = timezone.now().date()

    total_appointments = Appointment.objects.filter(doctor=doctor).count()
    new_appointments = Appointment.objects.filter(doctor=doctor, status__iexact="Pending").count()
    approved = Appointment.objects.filter(doctor=doctor, status__iexact="Approved").count()
    cancelled = Appointment.objects.filter(doctor=doctor, status__iexact="Cancelled").count()
    total_added_patients = AddedPatient.objects.filter(doctor=doctor).count()
    next_appointments = Appointment.objects.filter(
        doctor=doctor,
        date__gte=today,
    ).exclude(status__iexact="Cancelled").order_by("date", "time")[:5]
    recent_patients = AddedPatient.objects.filter(doctor=doctor).order_by("-created_at")[:5]

    context = {
        "total_appointments": total_appointments,
        "new_appointments": new_appointments,
        "approved": approved,
        "cancelled": cancelled,
        "total_added_patients": total_added_patients,
        "next_appointments": next_appointments,
        "recent_patients": recent_patients,
    }
    return render(request, "doctor/doctor_dashboard.html", context)



@login_required
@user_passes_test(is_doctor)
def doctor_appointments(request):
    search = request.GET.get("search", "").strip()
    status = request.GET.get("status", "").strip()

    appointments = Appointment.objects.filter(doctor=request.user)

    if search:
        appointments = appointments.filter(
            Q(patient_name__icontains=search)
            | Q(patient_mobile__icontains=search)
            | Q(symptoms__icontains=search)
        )

    if status:
        appointments = appointments.filter(status__iexact=status)

    appointments = appointments.order_by("date", "time")
    return render(
        request,
        "doctor/doctor_appointments.html",
        {"appointments": appointments, "search": search, "status": status},
    )


@login_required
@user_passes_test(is_doctor)
@require_POST
def approve_appointment(request, pk):
    app = get_object_or_404(Appointment, pk=pk, doctor=request.user)
    if app.status == "Cancelled":
        messages.error(request, "Cancelled appointments cannot be approved.")
    elif app.status == "Approved":
        messages.info(request, "This appointment is already approved.")
    else:
        app.status = "Approved"
        app.save(update_fields=["status"])
        AddedPatient.objects.get_or_create(
            doctor=request.user,
            mobile=app.patient_mobile,
            defaults={
                "name": app.patient_name,
                "age": 0,
                "disease": app.symptoms or ""
            }
        )
        messages.success(request, "Appointment approved successfully.")
    return redirect(request.POST.get("next") or "doctor:doctor_appointments")


@login_required
@user_passes_test(is_doctor)
@require_POST
def cancel_appointment(request, pk):
    app = get_object_or_404(Appointment, pk=pk, doctor=request.user)
    if app.status == "Cancelled":
        messages.info(request, "This appointment is already cancelled.")
    else:
        app.status = "Cancelled"
        app.save(update_fields=["status"])
        messages.success(request, "Appointment cancelled successfully.")
    return redirect(request.POST.get("next") or "doctor:doctor_appointments")


@login_required
@user_passes_test(is_doctor)
def doctor_patients(request):
    search = request.GET.get("search", "").strip()
   
    patients = AddedPatient.objects.filter(
    doctor=request.user,
    status="Active")

    if search:
        patients = patients.filter(
            Q(name__icontains=search)
            | Q(mobile__icontains=search)
            | Q(disease__icontains=search)
        )

    patients = patients.order_by("-created_at")
    return render(request, "doctor/doctor_patients.html", {"patients": patients, "search": search})


@login_required
@user_passes_test(is_doctor)
def add_patient(request):
    if request.method == "POST":
        AddedPatient.objects.create(
            doctor=request.user,
            name=request.POST["name"],
            mobile=request.POST["mobile"],
            age=request.POST.get("age"),
            disease=request.POST.get("disease", "")
        )
        messages.success(request, "Patient added successfully!")
        return redirect("doctor:doctor_patients")

    return render(request, "doctor/add_patient.html")


@login_required
@user_passes_test(is_doctor)
def edit_patient(request, pk):
    patient = get_object_or_404(AddedPatient, pk=pk, doctor=request.user)

    if request.method == "POST":
        patient.name = request.POST["name"]
        patient.mobile = request.POST["mobile"]
        patient.age = request.POST["age"]
        patient.disease = request.POST["disease"]
        patient.save()

        messages.success(request, "Patient updated successfully!")
        return redirect("doctor:doctor_patients")

    return render(request, "doctor/edit_patient.html", {"patient": patient})



@login_required
@user_passes_test(is_doctor)
def search_patient(request):
    query = request.GET.get("q", "").strip()
    filters = Q(name__icontains=query) | Q(mobile__icontains=query)
      

    if query.isdigit():
        filters |= Q(pk=int(query))

    patients = AddedPatient.objects.filter(
    doctor=request.user,
    status="Completed"
    ).filter(filters) 
    return render(request, "doctor/search_patient.html", {"patients": patients, "query": query})


@login_required
@user_passes_test(is_doctor)
def consultation_complete(request, patient_id):

    patient = get_object_or_404(
        AddedPatient,
        id=patient_id,
        doctor=request.user
    )

    return redirect(
        f"/billing/generate/?patient_id={patient.id}"
    )


@login_required
@user_passes_test(is_doctor)
def search_appointment(request):
    query = request.GET.get("q", "").strip()
    filters = Q(patient_name__icontains=query) | Q(patient_mobile__icontains=query)

    if query.isdigit():
        filters |= Q(pk=int(query))

    for date_format in ("%Y-%m-%d", "%d-%m-%Y"):
        try:
            parsed_date = datetime.strptime(query, date_format).date()
        except ValueError:
            continue
        filters |= Q(date=parsed_date)
        break

    appointments = Appointment.objects.filter(filters, doctor=request.user)
    return render(request, "doctor/search_appointment.html", {"appointments": appointments, "query": query})


@login_required
@user_passes_test(is_doctor)
def report_appointments(request):
    # Show all appointments by default
    apps = Appointment.objects.filter(doctor=request.user).order_by("-date", "-time")
    
    # Check if export is requested
    export_format = request.POST.get("export_format")
    if export_format == "excel":
        return export_appointments_excel(request)
    elif export_format == "pdf":
        return export_appointments_pdf(request)
    
    # If filters are applied, filter by date range
    if request.method == "POST":
        print(request.POST)
        start = request.POST.get("start")
        end = request.POST.get("end")
        if start and end:
            apps = apps.filter(
                date__range=[start, end]
            ).order_by("-date", "-time")
            print("After:", apps.count())
            print(apps.values_list("date", flat=True))
    
    return render(request, "doctor/report_appointments.html", {"appointments": apps})


@login_required
@user_passes_test(is_doctor)
def report_patients(request):
    # Show all patients by default
    patients = AddedPatient.objects.filter(doctor=request.user).order_by("-created_at")
    
    # Check if export is requested
    export_format = request.POST.get("export_format")
    if export_format == "excel":
        return export_patients_excel(request)
    elif export_format == "pdf":
        return export_patients_pdf(request)
    
    # If filters are applied, filter by date range
    if request.method == "POST":
        start = request.POST.get("start")
        end = request.POST.get("end")
        if start and end:
            patients = patients.filter(
                created_at__date__range=[start, end]
            ).order_by("-created_at")
    
    return render(request, "doctor/report_patients.html", {"patients": patients})




@login_required
@user_passes_test(is_doctor)
def doctor_profile(request):
    doctor = request.user

    if request.method == "POST":
        doctor.education = request.POST.get("education")
        doctor.experience = request.POST.get("experience")
        doctor.bio = request.POST.get("bio")
        doctor.fees = request.POST.get("fees")

        specialization_id = request.POST.get("specialization")
        if specialization_id:
            from accounts.models import Specialization
            doctor.specialization = Specialization.objects.get(id=specialization_id)

        doctor.save()
        messages.success(request, "Profile updated successfully!")
        return redirect("doctor:doctor_profile")

    from accounts.models import Specialization
    specializations = Specialization.objects.all()

    return render(request, "doctor/doctor_profile.html", {
        "doctor": doctor,
        "specializations": specializations
    })


@login_required
@user_passes_test(is_doctor)
def doctor_change_password(request):
    if request.method == "POST":
        old = request.POST["old_password"]
        new = request.POST["new_password"]

        if not request.user.check_password(old):
            messages.error(request, "Old password is incorrect")
            return redirect("doctor:doctor_change_password")

        request.user.set_password(new)
        request.user.save()
        messages.success(request, "Password changed successfully!")
        return redirect("accounts:login")

    return render(request, "doctor/doctor_change_password.html")


# ================== EXPORT FUNCTIONS ==================

@login_required
@user_passes_test(is_doctor)
def export_appointments_excel(request):
    """Export appointments to Excel"""
    appointments = Appointment.objects.filter(doctor=request.user).order_by("-date", "-time")
    
    if request.method == "POST":
        start = request.POST.get("start")
        end = request.POST.get("end")
        if start and end:
            appointments = appointments.filter(date__range=[start, end])
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appointments"
    
    # Add header
    headers = ["Patient Name", "Mobile", "Date", "Time", "Symptoms", "Status"]
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="1570A6", end_color="1570A6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for appt in appointments:
        ws.append([
            appt.patient_name,
            appt.patient_mobile,
            appt.date.strftime("%d-%m-%Y"),
            str(appt.time),
            appt.symptoms or "",
            appt.status
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="appointments_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(is_doctor)
def export_appointments_pdf(request):
    """Export appointments to PDF"""
    appointments = Appointment.objects.filter(doctor=request.user).order_by("-date", "-time")
    
    if request.method == "POST":
        start = request.POST.get("start")
        end = request.POST.get("end")
        if start and end:
            appointments = appointments.filter(date__range=[start, end])
    
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="appointments_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1570A6'),
        spaceAfter=30,
        alignment=1
    )
    elements.append(Paragraph(f"Sancheti Hospital - Appointment Report", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Create table data
    data = [["Patient", "Mobile", "Date", "Time", "Symptoms", "Status"]]
    for appt in appointments:
        data.append([
            appt.patient_name,
            appt.patient_mobile,
            appt.date.strftime("%d-%m-%Y"),
            str(appt.time),
            (appt.symptoms or "")[:30],
            appt.status
        ])
    
    # Create table
    table = Table(data, colWidths=[1.2*inch, 1*inch, 1*inch, 0.9*inch, 1.5*inch, 0.9*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1570A6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return response


@login_required
@user_passes_test(is_doctor)
def export_patients_excel(request):
    """Export patients to Excel"""
    patients = AddedPatient.objects.filter(doctor=request.user).order_by("-created_at")
    
    if request.method == "POST":
        start = request.POST.get("start")
        end = request.POST.get("end")
        if start and end:
            patients = patients.filter(created_at__date__range=[start, end])
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"
    
    # Add header
    headers = ["Patient Name", "Mobile", "Age", "Disease", "Created Date"]
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="1570A6", end_color="1570A6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for patient in patients:
        ws.append([
            patient.name,
            patient.mobile,
            patient.age or "",
            patient.disease or "",
            patient.created_at.strftime("%d-%m-%Y")
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="patients_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@user_passes_test(is_doctor)
def export_patients_pdf(request):
    """Export patients to PDF"""
    patients = AddedPatient.objects.filter(doctor=request.user).order_by("-created_at")
    
    if request.method == "POST":
        start = request.POST.get("start")
        end = request.POST.get("end")
        if start and end:
            patients = patients.filter(created_at__date__range=[start, end])
    
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="patients_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Add title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1570A6'),
        spaceAfter=30,
        alignment=1
    )
    elements.append(Paragraph(f"Sancheti Hospital - Patient Report", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Create table data
    data = [["Patient Name", "Mobile", "Age", "Disease", "Created Date"]]
    for patient in patients:
        data.append([
            patient.name,
            patient.mobile,
            str(patient.age) if patient.age else "",
            patient.disease or "",
            patient.created_at.strftime("%d-%m-%Y")
        ])
    
    # Create table
    table = Table(data, colWidths=[1.5*inch, 1.2*inch, 0.8*inch, 1.5*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1570A6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return response
